import openpyxl
import sys
import os
from typing import List, Tuple

# ANSI color codes
RED = '\033[91m'
GREEN = '\033[92m'
YELLOW = '\033[93m'
BLUE = '\033[94m'
MAGENTA = '\033[95m'
CYAN = '\033[96m'
RESET = '\033[0m'
BOLD = '\033[1m'

def get_cell_name(row: int, col: int) -> str:
    """Convert row/col numbers to Excel cell name (e.g., A1, B2)."""
    col_letter = openpyxl.utils.get_column_letter(col)
    return f"{col_letter}{row}"

def compare_xlsx_files(file1: str, file2: str) -> Tuple[bool, List[dict]]:
    """Compare two xlsx files and return (has_diff, list_of_diffs)."""
    diffs = []

    try:
        # Load workbooks in read-only mode for better performance
        wb1 = openpyxl.load_workbook(file1, read_only=True, data_only=True)
        wb2 = openpyxl.load_workbook(file2, read_only=True, data_only=True)

        # Compare number of sheets
        if wb1.sheetnames != wb2.sheetnames:
            diffs.append({
                'type': 'sheets',
                'old': wb1.sheetnames,
                'new': wb2.sheetnames
            })

        # Compare each sheet
        all_sheets = set(wb1.sheetnames) | set(wb2.sheetnames)
        for sheet_name in all_sheets:
            if sheet_name not in wb1.sheetnames:
                diffs.append({
                    'type': 'sheet_added',
                    'sheet': sheet_name
                })
                continue

            if sheet_name not in wb2.sheetnames:
                diffs.append({
                    'type': 'sheet_removed',
                    'sheet': sheet_name
                })
                continue

            ws1 = wb1[sheet_name]
            ws2 = wb2[sheet_name]

            # In read-only mode, we need to iterate through rows properly
            # Convert worksheets to lists of rows
            rows1 = list(ws1.rows)
            rows2 = list(ws2.rows)

            max_row = max(len(rows1), len(rows2))

            # Compare dimensions
            ws1_cols = max((len(row) for row in rows1), default=0)
            ws2_cols = max((len(row) for row in rows2), default=0)

            if len(rows1) != len(rows2) or ws1_cols != ws2_cols:
                diffs.append({
                    'type': 'dimensions',
                    'sheet': sheet_name,
                    'old': (len(rows1), ws1_cols),
                    'new': (len(rows2), ws2_cols)
                })

            # Compare cell values
            for row_idx in range(max_row):
                row1 = rows1[row_idx] if row_idx < len(rows1) else []
                row2 = rows2[row_idx] if row_idx < len(rows2) else []

                max_col = max(len(row1), len(row2))

                for col_idx in range(max_col):
                    val1 = row1[col_idx].value if col_idx < len(row1) else None
                    val2 = row2[col_idx].value if col_idx < len(row2) else None

                    if val1 != val2:
                        diffs.append({
                            'type': 'cell',
                            'sheet': sheet_name,
                            'cell': get_cell_name(row_idx + 1, col_idx + 1),
                            'row': row_idx + 1,
                            'col': col_idx + 1,
                            'old': val1,
                            'new': val2
                        })

        return len(diffs) > 0, diffs
    except Exception as e:
        print(f"{RED}Error comparing files: {e}{RESET}")
        import traceback
        traceback.print_exc()
        return True, []  # Assume different if comparison fails

def write_github_output(key: str, value: str):
    """Write to GITHUB_OUTPUT if running in GitHub Actions, otherwise skip."""
    github_output = os.environ.get('GITHUB_OUTPUT')
    if github_output:
        with open(github_output, 'a') as f:
            f.write(f"{key}={value}\n")

def print_diffs(diffs: List[dict]):
    """Print differences in a colorized format."""
    if not diffs:
        return

    print(f"\n{BOLD}{CYAN}{'='*80}{RESET}")
    print(f"{BOLD}{CYAN}Differences Found:{RESET}")
    print(f"{BOLD}{CYAN}{'='*80}{RESET}\n")

    current_sheet = None

    for diff in diffs:
        if diff['type'] == 'sheets':
            print(f"{YELLOW}Sheet names differ:{RESET}")
            print(f"  {RED}- Old: {diff['old']}{RESET}")
            print(f"  {GREEN}+ New: {diff['new']}{RESET}\n")

        elif diff['type'] == 'sheet_added':
            print(f"{GREEN}+ Sheet added: {diff['sheet']}{RESET}\n")

        elif diff['type'] == 'sheet_removed':
            print(f"{RED}- Sheet removed: {diff['sheet']}{RESET}\n")

        elif diff['type'] == 'dimensions':
            print(f"{YELLOW}Sheet '{diff['sheet']}' dimensions changed:{RESET}")
            print(f"  {RED}- Old: {diff['old'][0]} rows x {diff['old'][1]} cols{RESET}")
            print(f"  {GREEN}+ New: {diff['new'][0]} rows x {diff['new'][1]} cols{RESET}\n")

        elif diff['type'] == 'cell':
            # Print sheet header if we're in a new sheet
            if current_sheet != diff['sheet']:
                current_sheet = diff['sheet']
                print(f"{BOLD}{MAGENTA}Sheet: {current_sheet}{RESET}")
                print(f"{MAGENTA}{'-'*80}{RESET}")

            cell_ref = f"{diff['cell']}"
            old_val = repr(diff['old']) if diff['old'] is not None else 'None'
            new_val = repr(diff['new']) if diff['new'] is not None else 'None'

            print(f"  {BLUE}Cell {cell_ref}:{RESET}")
            print(f"    {RED}- {old_val}{RESET}")
            print(f"    {GREEN}+ {new_val}{RESET}")

    print(f"\n{BOLD}{CYAN}{'='*80}{RESET}")
    print(f"{BOLD}{CYAN}Total differences: {len(diffs)}{RESET}")
    print(f"{BOLD}{CYAN}{'='*80}{RESET}\n")

# Check if original file exists
original_file = "csv/merge.xlsx"
new_file = "merge_new.xlsx"

if not os.path.exists(original_file):
    print("Original file does not exist, will create new one")
    write_github_output("has_diff", "true")
    sys.exit(0)

# Compare files
has_diff, diffs = compare_xlsx_files(original_file, new_file)

if has_diff:
    print(f"{GREEN}Files differ - will update merge.xlsx{RESET}")
    print_diffs(diffs)
    write_github_output("has_diff", "true")
else:
    print(f"{CYAN}Files are identical - no update needed{RESET}")
    write_github_output("has_diff", "false")
